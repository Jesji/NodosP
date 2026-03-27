# Funciones transformacion
import unicodedata
import pandas as pd
import openpyxl 
import glob, os, re, shutil, time, json
import numpy as np
from pathlib import Path


def importar_munis(file_muni):
    df = pd.read_excel(file_muni, skiprows=3)
    columns_to_drop = list(['Estatus']) + list(df.columns[-9:])
    df = df.drop(columns=columns_to_drop)
    df = formatear_texto_columnas(df)
    return df

def formatear_texto(name):
    """
    Funcion para quitar acentos de una frase, pasar a MAYUSCULA y quitar puntos
    
    :param name: Es el nombre de las columnas del df
    """
    # se crea un diccionario para los valores en unicode de acento y dieresis y punto
    trans_tab = dict.fromkeys(map(ord, u'\u0301\u0308.'), None) # ord pasa de unicode a un numero
    if isinstance(name,(str,list, tuple,np.ndarray)):
          sin_acento = unicodedata.normalize('NFKD', name).translate(trans_tab)
          sin_acento = sin_acento.upper() # pasar a mayuscula
          sin_acento = sin_acento.strip() # quitar espacios en blanco al inicio y al final
          #print(sin_acento)
          return sin_acento
    else:
          return name

def formatear_texto_columnas(df:pd.DataFrame) -> pd.DataFrame:
     for col in df.columns:
          df[col] = df[col].map(formatear_texto)
     print('\nTexto de columnas formateado')
     return df


def find_header(file,keyword):
        """
        Funcion para encontrar header en donde aparece cierta palabra
        """
        wb = openpyxl.load_workbook(file)
        sheets = wb.sheetnames  # list sheets
        sheet = wb[sheets[0]]   # choose first sheet 
        for line_idx, str_line in enumerate(sheet.iter_rows(values_only=True)):
            # leer linea
            #print(str_line)
            if any(keyword.lower() in str(cell).lower() for cell in str_line if cell is not None):
                return line_idx
        wb.close()
        raise ValueError(f'No se encontro encabezado en {file}')

def read_file(file):
     line_header = find_header(file,'SISTEMA')
     df=pd.read_excel(file, skiprows=line_header, header=0)
     return df

def read_concat(folder_path):
    list_files = glob.glob(os.path.join(folder_path,'*.xlsx'))
    dataframes = [] # lista de df
     
    for file in list_files : # loop para la lista de archivos
        if 'Sistema' in file:
            try:
                if any(char.isdigit() for char in file): # en los caracteres numericos 
                    searchre = re.search(r'\d{4}-\d{2}-\d{2}', file) # buscar el formato yyyy-mm-dd
                    fecha = searchre.group() # extraer resultado de busqueda
                    print(f"Leyendo archivo del {fecha}")
                df = read_file(file)
                df.columns = [formatear_texto(c) for c in df.columns]

                df['Date'] = pd.to_datetime(fecha)  # crear columna fecha y que tome valor de fecha en el nonmbre
               
                dataframes.append(df)   # agregar df a la lista     
            except Exception as e:
                print(f'Error {e} en file {file}')
        else: continue
    if dataframes:
        df_tot = pd.concat(dataframes, ignore_index=True)
        print(f'\nTotal filas en df_tot: {len(df_tot)}')
          
    return df_tot 

def reduce_columns(df:pd.DataFrame) -> pd.DataFrame: 
    """
    Reduce las columnas del df de los catalogos de NodosP
    """ 
    columns_to_drop = [] # crear lista de columnas para quitar 
    for col in df.columns:
        if 'MODELADA' in col or 'TRANSMISION' in col: # si contiene la raiz modelada
            columns_to_drop.append(col)
    # quitar las columnas con modelada y las ultimas 3 
    #columns_to_drop = columns_to_drop + list(df.columns[-3:]) 
    df_reduced = df.drop(columns=columns_to_drop)
    # Rename certain columns to match the municipios df
    df_reduced = df_reduced.rename(columns={
        'LOCALIDAD':'NOM_LOC',
        'CLAVE DE ENTIDAD FEDERATIVA (INEGI)': 'CVE_ENT',
        'ENTIDAD FEDERATIVA (INEGI)': 'NOM_ENT',
        'CLAVE DE MUNICIPIO (INEGI)' : 'CVE_MUN',         
        'MUNICIPIO (INEGI)' : 'NOM_MUN' ,
        'ESTADO':'NOM_EST',
        'CENTRO DE CONTROL REGIONAL': 'CCR'  ,
        'NIVEL DE TENSION (KV)':'TENSION'
    })
    print('Columnas del df reducidas')
    return df_reduced

def sacar_de_lista(celda, default = pd.NA):
    """ 
    Esta funcion revisa si el valor de una celda es una lista de un elemento lo retorna
    como string o escalar. Si tiene mas de un elemento conserva la lista. Si ya deporsi 
    era escalar, deja el escalar. 
    """
    
    # checar tipo de dato celda
    if isinstance(celda, (list, tuple, np.ndarray)):
        # obtener valores de la celda
        valores = [v for v in celda if pd.notna(v)]

        if len(valores) == 0: # lista esta vacia
            return default
        if len(valores) == 1: # lista solo tiene un elemento 
            return valores[0]
        else:   
            return valores # lista tiene mas de un elemento
        
    # # caso la celda es nan
    # if pd.isna(celda):
    #     return default
    
    else:
        return celda  # si celda no es lista o tupla

def agrupar_por_clave(df:pd.DataFrame)->pd.DataFrame:
    """ 
    Esta funcion crea un df que agrupa los valores del df por CLAVE del nodo
    considerando las columnas y agregaciones indicadas en el dictionary
    """
    dict_agg = {
        'NOM_LOC': ('NOM_LOC', 'unique'),
        'NOM_EST': ('NOM_EST', 'unique'),
        'NOM_ENT': ('NOM_ENT', 'unique'),
        'CVE_ENT': ('CVE_ENT', 'unique'),
        'NOM_MUN': ('NOM_MUN', 'unique'),
        'CVE_MUN': ('CVE_MUN', 'unique'),
        'SISTEMA': ('SISTEMA', 'unique'),
        'CCR': ('CCR', 'unique'),
        'TENSION': ('TENSION', 'last'),
        'Num_NOM_EST': ('NOM_EST', 'nunique'),
        'Num_NOM_LOC': ('NOM_LOC', 'nunique'),
        'Num_NOM_ENT': ('NOM_ENT', 'nunique'),
        'Num_CVE_ENT': ('CVE_ENT', 'nunique'),
        'Num_NOM_MUN': ('NOM_MUN', 'nunique'),
        'Num_CVE_MUN': ('CVE_MUN', 'nunique'),
    }
    # filtrar dict si la columna esta en el df 
    dict_aggf = {k:v for k,v in dict_agg.items() if v[0] in df.columns}

    df_agrupado = df.copy().groupby(by=['CLAVE']).agg(**dict_aggf)
    
    #cols_to_keep = ["NOM_EST","NOM_LOC","NOM_ENT","CVE_ENT",
    #    "NOM_MUN","CVE_MUN",'SISTEMA','CCR','TENSION']

    for col in df_agrupado.columns:
        df_agrupado[col] = df_agrupado[col].apply(sacar_de_lista)
    print("El df ha sido agrupado por CLAVE")
    return df_agrupado

# Manejar no aplica 
def drop_noaplica(df:pd.DataFrame):
    mask = (df=='NO APLICA').any(axis=1)
    df_no_aplica = df[mask]
    df_clean = df[~mask]
    return df_clean, df_no_aplica

# Manejar all nan 
def drop_all_nan(df:pd.DataFrame):
    mask = df.isna().all(axis=1)
    df_all_na = df[mask]
    df_clean = df[~mask]
    return df_clean, df_all_na

def limpiar_string_array(texto):
    # Si no es texto (es NaN o None), regresamos vacío o el mismo valor
    if not isinstance(texto, str):
        return texto
    
    # Buscamos lo que esté entre comillas simples dentro de los corchetes
    match = re.search(r"\[['\"](.+?)['\"]\]", texto)
    if match:
        return match.group(1)
    
    # Si no tiene formato de lista, devolvemos el texto original limpio
    return texto.strip()

def agrupar_mun(df:pd.DataFrame):
    """
    Agrupa el df del catalogo de poblaciones del inegi por cve_ent y cve_mun
    """
    df_agrupado = (df.groupby(['CVE_ENT','CVE_MUN'], as_index=False)
                   .agg(
                       CVE_ENT = ('CVE_ENT', 'unique'),
                       NOM_ENT = ('NOM_ENT', 'unique'),
                       CVE_MUN = ('CVE_MUN', 'unique'),
                       NOM_MUN = ('NOM_MUN', 'unique'),
                       LAT_DECIMAL = ('LAT_DECIMAL','mean'),
                       LON_DECIMAL = ('LON_DECIMAL','mean')
                   ))
    for col in df_agrupado.columns:
        df_agrupado[col] = df_agrupado[col].apply(sacar_de_lista)
    df_agrupado['NOM_ENT'] = df_agrupado['NOM_ENT'].astype(str).apply(limpiar_string_array)
    df_agrupado['NOM_MUN'] = df_agrupado['NOM_MUN'].astype(str).apply(limpiar_string_array)
    df_agrupado = df_agrupado.drop_duplicates()
    # Obtener dictioary cve_mun a nom_mun
    dict_cvemun_nommun_canonico = df_agrupado.groupby('CVE_ENT').apply(
         lambda x:dict(zip(x['CVE_MUN'],x['NOM_MUN']))
     ).to_dict()
    dict_nommun_cvemun_canon = df_agrupado.groupby('CVE_ENT').apply(
         lambda x:dict(zip(x['NOM_MUN'],x['CVE_MUN']))
     ).to_dict()
    return df_agrupado , dict_nommun_cvemun_canon, dict_cvemun_nommun_canonico
    #return df_agrupado 

def quitar_est_incorre(row):
    """
    Funcion que quita los valores NOMEST diferentes de NOMENT, y en los 
    casos cdmx deja los valores que contienen "CD." Y en el caso de 
    EDOMEX deja los casos que contienen EDO o ESTADO.
    
    :param row: Description
    """
    nomest = row['NOM_EST']
    noment = row['NOM_ENT']
    # si el nomest es igual al noment
    if nomest in noment:
        return nomest
    # caso ciudad de mexico
    if noment == 'CIUDAD DE MEXICO' and 'CD' in nomest :
        return nomest 
    # caso estado de mexico
    if noment == 'MEXICO' and ('EDO' in nomest ) or ('ESTADO' in nomest) :
        return nomest
    else:
        return pd.NA

# ESta funcion esta bien pero es mas complicada porque usa apply NO SE RECOMIENDA
def map_nomest_cveent_noment(row,dict_cve_ent,dict_cveent_noment):
    cveent = row['NOM_ENT']
    nomest = row['NOM_EST']
    noment = row['NOM_ENT']
    # caso que hay nomest pero no cveent
    if pd.isna(cveent) and pd.notna(nomest):
        cveent = dict_cve_ent.get(nomest,pd.NA)
    if pd.isna(noment) and pd.notna(cveent):
        noment = dict_cveent_noment.get(cveent,pd.NA)
    return pd.Series({'CVE_ENT':cveent, 'NOM_ENT': noment})

    
def fix_missing_cve_ent(df:pd.DataFrame):
    """
    Esta funcion produce dictionarios para mapear de NOM_EST a CVE_ENT y luego de CVE_ENT a NOM_ENT
    
    :param df: Description
    :type df: pd.DataFrame
    """
    df_u_exploded = (df
        .copy()
        # cuando nom_est es una lista, crea filas para cada posible nom_est
        .explode("NOM_EST") 
        .groupby(['CVE_ENT','NOM_ENT'],as_index=False).agg(
            NOM_EST = ('NOM_EST','unique')
        )                   
    )
    # expandir por nom_est
    df_nom_est_inconsis = df_u_exploded.explode(column='NOM_EST').dropna()
    # quitar los no aplica
    df_nom_est_inconsis = df_nom_est_inconsis[df_nom_est_inconsis['NOM_ENT']!='NO APLICA'].drop_duplicates()
    # quita NOMEST que no coincide con NOMENT y tratar casos CDMX y EDOMEX
    df_nom_est_inconsis['NOM_EST']= df_nom_est_inconsis.apply(quitar_est_incorre,axis=1)
    # get dict de NOM_EST a CVE_ENT
    df_inconsi = df_nom_est_inconsis.dropna().set_index('NOM_EST')
    dict_nomest_cveent = df_inconsi['CVE_ENT'].to_dict()
    # get dict CVE_ENT a NOM_ENT
    df_cve_ent = df_nom_est_inconsis.dropna().set_index('CVE_ENT')
    dict_cveent_noment = df_cve_ent['NOM_ENT'].to_dict()

    df_fixed = df.copy()
    # map nomest a cveent
    mask_cve = df_fixed['CVE_ENT'].isna() & df_fixed['NOM_EST'].notna()
    df_fixed.loc[mask_cve, 'CVE_ENT'] = (df_fixed.loc[mask_cve, 'NOM_EST'].map(dict_nomest_cveent))

    # MAP CVEENT A NOMENT
    mask_noment = df_fixed['NOM_ENT'].isna() & df_fixed['CVE_ENT'].notna()
    df_fixed.loc[mask_noment,'NOM_ENT'] =( df_fixed.loc[mask_noment, 'CVE_ENT'].map(dict_cveent_noment) )

    return df_fixed

def map_nomloc_cvemun(row,dict_nomloc_cvemun,dict_cvemun_nommun,dict_nommun_cvemun_canon,dict_cvemun_nommun_canon):
#def map_nomloc_cvemun(row,dict_nomloc_cvemun,dict_cvemun_nommun):
    cveent = row['CVE_ENT']
    cvemun = row['CVE_MUN']
    nomloc = row['NOM_LOC']
    nommun = row['NOM_MUN']
    # si hay NOM_LOC y no hay CVE_MUN
    if pd.notna(nomloc) & pd.isna(cvemun):
        cvemun = dict_nomloc_cvemun.get(cveent,{}).get(nomloc,pd.NA)
        if pd.isna(cvemun):
            for key,val in dict_nomloc_cvemun.get(cveent,{}).items():
                if (key in nomloc) or (nomloc in key):
                    cvemun = val
            if pd.isna(cvemun):
                for key,val in dict_nommun_cvemun_canon.get(cveent,{}).items():
                    if (key in nomloc) or (nomloc in key):
                        cvemun = val           
    if pd.notna(cvemun) & pd.isna(nommun):
        nommun = dict_cvemun_nommun.get(cveent,{}).get(cvemun,pd.NA)
        if pd.isna(nommun):
            nommun = dict_cvemun_nommun_canon.get(cveent,{}).get(cvemun,pd.NA)

    return pd.Series({'CVE_MUN': cvemun, 'NOM_MUN':nommun})


def fix_missing_cvemun(df:pd.DataFrame, dict_nommun_cvemun_canon:dict, dict_cvemun_nommun_canon:dict)->pd.DataFrame:
    """
    Funcion que modifica el df para rellenar valores faltantes de CVEMUN a partir de valores de NOM_LOC
    
    :param df: Description
    :type df: pd.DataFrame
    :return: Description
    :rtype: DataFrame
    """
    df_ag_nomloc = df.copy().explode(['NOM_LOC']).groupby(['CVE_ENT','NOM_LOC'], as_index =False).agg(
        Num_NOM_LOC = ('NOM_LOC','nunique'),
        NOM_MUN = ('NOM_MUN','unique'),
        CVE_MUN = ('CVE_MUN','unique'),
        #Num_NOM_MUN = ('NOM_MUN','nunique') # -> ESTO arroja error unhashable type: 'list'
    )

    # Quitar combinaciones duplicadas de CVE_, CVE_MUN, NOM_MUN, 
    df_cve_nom_mun = (df_ag_nomloc
                      .explode(['CVE_MUN','NOM_MUN'])
                      .drop_duplicates(subset=['CVE_ENT','CVE_MUN','NOM_MUN'])
                      .dropna()
                      .sort_values(['CVE_ENT','CVE_MUN']))
 
    # Obtener dictionario nomloc a cvemun
    dict_nomloc_cvemun = df_cve_nom_mun.groupby('CVE_ENT').apply(
        # esta funcion lambda crea un dict del tipo {cve_ent:{nom_loc:cve_mun,nom_loc:cve_mun,...}}
        lambda x: dict(zip(x['NOM_LOC'],x['CVE_MUN']))
    ).to_dict()

    # Obtener dictioary cve_mun a nom_mun
    dict_cvemun_nommun = df_cve_nom_mun.groupby('CVE_ENT').apply(
        lambda x: dict(zip(x['CVE_MUN'],x['NOM_MUN']))
    ).to_dict()

    # copiar df original
    df_fixed = df.copy()
    # Sustituir valores donde CVE_MUN es na USANDO APPLY 
    cvemun_na_mask = df_fixed['NOM_LOC'].notna() & df_fixed['CVE_MUN'].isna()
    df_fixed.loc[cvemun_na_mask,['CVE_MUN','NOM_MUN']] = df_fixed.loc[cvemun_na_mask].apply(
        map_nomloc_cvemun, axis=1, dict_nomloc_cvemun=dict_nomloc_cvemun,
        dict_cvemun_nommun=dict_cvemun_nommun, 
        dict_nommun_cvemun_canon = dict_nommun_cvemun_canon,
        dict_cvemun_nommun_canon = dict_cvemun_nommun_canon)
    
    return df_fixed

def keep_cols_clean(df:pd.DataFrame):
    columns_to_keep = ['CLAVE','CVE_ENT',
                    'NOM_ENT','CVE_MUN',
                    'NOM_MUN','SISTEMA','CCR','TENSION']
    df_clean = df[columns_to_keep]
    # df_clean = df_clean.reset_index()
    # quitar filas para las que las 4 cols de interes son na
    filter = df_clean[['CVE_ENT','NOM_ENT','CVE_MUN',
                    'NOM_MUN']].isna().all(axis=1)
    df_clean = df_clean[~filter]
    # asignar tipos de datos
    dict_types = {
        'CLAVE':'str', 'CVE_ENT':'int', 'NOM_ENT':'str', 'CVE_MUN':'int',
        'NOM_MUN':'str', 'SISTEMA':'str', 'CCR':'str',
        'TENSION':'str'
    }
    df_clean = df_clean.astype(dict_types)
    return df_clean

def mapear_faltantes_catalog(df_in:pd.DataFrame, df_ag:pd.DataFrame):
    # mapear cve_ent y cve_mun a los valores del df_agrupado
    ent_map = df_ag.set_index('CLAVE')['CVE_ENT']
    nom_ent_map = df_ag.set_index('CLAVE')['NOM_ENT']
    mun_map = df_ag.set_index('CLAVE')['CVE_MUN']
    nom_mun_map = df_ag.set_index('CLAVE')['NOM_MUN']
    # sustituir por la serie en el df_ag
    df = df_in.copy()
    df['CVE_ENT'] = df['CVE_ENT'].fillna(df['CLAVE'].map(ent_map))
    df['NOM_ENT'] = df['NOM_ENT'].fillna(df['CLAVE'].map(nom_ent_map))
    df['CVE_MUN'] = df['CVE_MUN'].fillna(df['CLAVE'].map(mun_map))
    df['NOM_MUN'] = df['NOM_MUN'].fillna(df['CLAVE'].map(nom_mun_map))

    return df

def quitar_cvemun_duplicados(df:pd.DataFrame):
    # filtrar filas con listas
    mask = df.map(lambda x: isinstance(x,(list))).any(axis=1) 
    df_list = df[mask]
    if len(df_list) > 0:
        # si hay listas en alguna celda
        df_fixed = (df.copy().reset_index()
                    .explode(['CVE_MUN','NOM_MUN'])
                    .drop_duplicates(subset=['CLAVE'],keep='last'))
    else: df_fixed =  df.copy()

    return df_fixed
    
def create_folder(folder_name):
    # download folder name
    download_folder = Path.cwd()/f"{folder_name}"
    # si el folder existe, eliminalo
    if download_folder.exists():
        shutil.rmtree(download_folder)
    download_folder.mkdir(parents=True, exist_ok=True)
    print(f"Folder {folder_name}/ creado")
    return download_folder

def imprimir_bienvenida():
    # Limpia la terminal 
    os.system('cls' if os.name == 'nt' else 'clear')
    
    # Arte ASCII de "NodosP"
    banner = """
    ============================================================
    ||                                                        ||
    ||   _   _           _           ____                     ||
    ||  | \ | | ___   __| | ___  ___|  _ \                    ||
    ||  |  \| |/ _ \ / _` |/ _ \/ __| |_) |                   ||
    ||  | |\  | (_) | (_| | (_) \__ \  __/                    ||
    ||  |_| \_|\___/ \__,_|\___/|___/_|                       ||
    ||                                                        ||
    ||        SISTEMA DE EXTRACCIÓN Y ANÁLISIS 2026           ||
    ============================================================
    """
    print(banner)
    time.sleep(1)

def delimitar_CCR(fila, dict_file):
    nom_ent = fila['NOM_ENT']
    nom_mun = fila['NOM_MUN']
    
    # read dictionary
    with open(dict_file, 'r') as f:
        delimitaciones = json.load(f)

    for ccr,condiciones in delimitaciones.items():
        if nom_ent in condiciones['estados_completos']:
            return ccr
        if nom_ent in condiciones['estados_parciales'].keys():
            for edo_par, definiciones in condiciones['estados_parciales'].items():
                incluir = definiciones['incluir']
                excluir = definiciones['excluir']
                if (incluir == 'ALL') & (nom_mun not in excluir):
                    return ccr
                if (excluir == 'ALL') & (nom_mun in incluir):
                    return ccr
